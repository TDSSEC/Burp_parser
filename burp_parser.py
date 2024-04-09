import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import argparse

def parse_burp_xml(xml_file, output_file):
    try:
        print(f"Opening XML file: {xml_file}")
        # Parse the XML file
        tree = ET.parse(xml_file)
        root = tree.getroot()

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write headers to the first row of the worksheet
        write_header(ws)

        # Write issues to the worksheet
        write_issues_to_xlsx(root, ws)

        # Create Excel table (ListObject) with specified style
        create_excel_table(ws, "A1:J100", "Table1", show_headers=True, table_style="TableStyleMedium2")

        # Save the workbook to the specified output file
        wb.save(output_file)

        print(f"XLSX file '{output_file}' successfully created with data and table.")
    except Exception as e:
        print(f"Error parsing XML file: {e}")
    finally:
        print("Closing XML file.")

def write_header(ws):
    # Define headers for the columns
    headers = ["Name", "Host", "IP", "Path", "Severity", "Confidence",
               "Issue Background", "Remediation Detail",
               "Vulnerability Classification", "Issue Details"]

    # Write headers to the first row of the worksheet
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

def write_issues_to_xlsx(root, ws):
    # Iterate over issues and write data to the worksheet
    row_index = 2  # Start writing data from the second row

    for issue in root.findall('.//issue'):
        try:
            # Extract data from each issue
            name = issue.find('name').text.strip() if issue.find('name') is not None else ''
            host = issue.find('host').text.strip() if issue.find('host') is not None else ''
            ip = issue.find('host').attrib.get('ip', '') if issue.find('host') is not None else ''
            path = issue.find('path').text.strip() if issue.find('path') is not None else ''
            severity = issue.find('severity').text.strip() if issue.find('severity') is not None else ''
            confidence = issue.find('confidence').text.strip() if issue.find('confidence') is not None else ''
            issue_background = issue.find('issueBackground').text.strip() if issue.find('issueBackground') is not None else ''
            remediation_detail = issue.find('remediationDetail').text.strip() if issue.find('remediationDetail') is not None else ''
            classification = issue.find('type').text.strip() if issue.find('type') is not None else ''
            issue_details = issue.find('issueDetail').text.strip() if issue.find('issueDetail') is not None else ''

            # Write data to the worksheet
            ws.cell(row=row_index, column=1, value=name)
            ws.cell(row=row_index, column=2, value=host)
            ws.cell(row=row_index, column=3, value=ip)
            ws.cell(row=row_index, column=4, value=path)
            ws.cell(row=row_index, column=5, value=severity)
            ws.cell(row=row_index, column=6, value=confidence)
            ws.cell(row=row_index, column=7, value=issue_background)
            ws.cell(row=row_index, column=8, value=remediation_detail)
            ws.cell(row=row_index, column=9, value=classification)
            ws.cell(row=row_index, column=10, value=issue_details)

            row_index += 1  # Move to the next row for the next issue
        except Exception as e:
            print(f"Error processing issue: {e}")

def create_excel_table(ws, ref, name, show_headers=True, table_style=None):
    # Create Excel table (ListObject) on the worksheet
    table = Table(displayName=name, ref=ref)

    # Set table style options
    if table_style:
        style = TableStyleInfo(name=table_style, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

def main():
    # Create argument parser
    parser = argparse.ArgumentParser(description='Parse Burp XML file and export to XLSX format with table.')
    parser.add_argument('xml_file', type=str, help='Path to Burp XML file')
    parser.add_argument('output_file', type=str, help='Output XLSX file name')
    args = parser.parse_args()

    # Parse the Burp XML file and write issues to XLSX with table
    print(f"Starting XML parsing and data extraction for '{args.xml_file}'...")
    parse_burp_xml(args.xml_file, args.output_file)
    print("Process completed.")

if __name__ == "__main__":
    main()
